import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ──────────────────────────────────────────────
# Helper functions
# ──────────────────────────────────────────────

def parse_deviation(val):
    """Parse '01h 32m - overtime' → (overtime_mins, undertime_mins)."""
    if pd.isna(val):
        return 0, 0
    m = re.search(r'(\d+)h (\d+)m - (undertime|overtime)', str(val), re.IGNORECASE)
    if m:
        h, mins = int(m.group(1)), int(m.group(2))
        total = h * 60 + mins
        return (total, 0) if m.group(3).lower() == 'overtime' else (0, total)
    return 0, 0

def parse_hm(val):
    """Parse '12:35' or '-1:23' → total minutes (signed)."""
    if pd.isna(val):
        return 0
    val_str = str(val).strip()
    if not val_str or val_str in ('0', 'nan'):
        return 0
    negative = val_str.startswith('-')
    val_str = val_str.lstrip('-')
    parts = val_str.split(':')
    if len(parts) >= 2:
        try:
            total = int(parts[0]) * 60 + int(parts[1])
            return -total if negative else total
        except ValueError:
            return 0
    return 0

def format_hm(total_mins, blank_if_zero=False):
    """
    Format signed minutes → plain text string like '12:35' or '-1:23'.
    Returns '' if blank_if_zero=True and value is 0.
    """
    if total_mins == 0:
        return '' if blank_if_zero else '0'
    sign = '-' if total_mins < 0 else ''
    total_mins = abs(int(total_mins))
    return f"{sign}{total_mins // 60}:{total_mins % 60:02d}"


def to_excel_text(df):
    """
    Write DataFrame to an xlsx bytes buffer.
    Every data cell is written as a plain TEXT string so Excel
    never auto-converts "1:06" → time or "-31:22" → #VALUE!
    """
    wb = Workbook()
    ws = wb.active

    # Header row
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Data rows – force every cell to Text number format
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = '@'          # Excel "Text" format
            cell.value = str(value) if value != '' else ''
            cell.alignment = Alignment(horizontal='left')

    # Auto-fit column widths roughly
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# Core processing
# ──────────────────────────────────────────────

def process_data(attendance_file, dump_file):
    # 1. Read Process Dump CSV
    df_dump = pd.read_csv(dump_file)
    opening_col = 'New Month - Excess or shortage hours'
    email_to_opening_mins = {}

    if 'Submitter Email' in df_dump.columns and opening_col in df_dump.columns:
        for _, row in df_dump.dropna(subset=['Submitter Email']).iterrows():
            email = str(row['Submitter Email']).strip()
            email_to_opening_mins[email] = parse_hm(row[opening_col])

    # 2. Read Attendance CSV and aggregate by Email + Store
    df_att = pd.read_csv(attendance_file)
    df_att[['OT_Mins', 'UT_Mins']] = df_att.apply(
        lambda r: pd.Series(parse_deviation(r.get('Deviation (Over time / Under time)'))),
        axis=1
    )
    df_att['Email'] = df_att['Email'].astype(str).str.strip()
    df_att['Store'] = df_att['Store'].astype(str).str.strip()

    df_grouped = (
        df_att
        .groupby(['Email', 'Store'], as_index=False)
        .agg({'OT_Mins': 'sum', 'UT_Mins': 'sum'})
    )

    # 3. Build output rows
    submission_date = datetime.today().strftime('%d-%m-%Y')
    results = []

    for _, row in df_grouped.iterrows():
        email = row['Email']
        if email == 'nan' or not email:
            continue

        store      = row['Store']
        ot_mins    = row['OT_Mins']
        ut_mins    = row['UT_Mins']
        opening    = email_to_opening_mins.get(email, 0)
        new_excess = opening + ot_mins - ut_mins

        # Label: Excess if >= 0, Shortage if negative
        label = 'Shortage' if new_excess < 0 else 'Excess'
        # The hours column is always shown as a positive (absolute) value
        abs_excess = abs(new_excess)

        results.append({
            'Submission Date':                                  submission_date,
            'Store Name':                                       store,
            'User Email':                                       email,
            'Last Month -  Excess or shortage hours-0100001':  format_hm(opening),
            'New Month - Total overtime-0100002':               format_hm(ot_mins,    blank_if_zero=True),
            'New Month - Total undertime-0100003':              format_hm(ut_mins,    blank_if_zero=True),
            'New Month - Excess or shortage hours-0100004':     format_hm(abs_excess),
            'Total is Excess or Shortage-0100005':              label,
        })

    return pd.DataFrame(results)


# ──────────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────────

st.set_page_config(page_title="Excess Hours Calculator", layout="centered", page_icon="⏱️")

st.title("⏱️ Excess Hours Calculator")
st.markdown(
    "Upload the current month's **Attendance CSV** and last month's **Process Dump CSV** "
    "to generate the Bulk Upload Excel for this month."
)

st.header("Step 1: Upload Files")
col1, col2 = st.columns(2)

with col1:
    attendance_file = st.file_uploader(
        "1️⃣ Attendance CSV  *(current month)*",
        type=["csv"],
        help="Daily attendance export from your system."
    )

with col2:
    dump_file = st.file_uploader(
        "2️⃣ Process Dump CSV  *(last month)*",
        type=["csv"],
        help="Monthly Excess Hours dump downloaded from your process module."
    )

if attendance_file and dump_file:
    st.header("Step 2: Generate Report")

    if st.button("⚡ Calculate Excess Hours", type="primary", use_container_width=True):
        with st.spinner("Processing…"):
            try:
                result_df = process_data(attendance_file, dump_file)

                st.success(f"✅ Successfully processed **{len(result_df)}** employee records!")

                st.subheader("Preview")
                st.dataframe(result_df, use_container_width=True, hide_index=True)

                # Excel download (text-formatted cells — no #VALUE! errors)
                excel_bytes = to_excel_text(result_df)
                file_name = f"Process_Bulk_Upload_{datetime.today().strftime('%Y_%m')}.xlsx"
                st.download_button(
                    label="📥 Download Bulk Upload Excel",
                    data=excel_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"Something went wrong: {e}")
